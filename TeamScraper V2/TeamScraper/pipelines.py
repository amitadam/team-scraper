# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import json
import os
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


class TeamscraperPipeline:
    def process_item(self, item, spider):
        return item


class IncrementalJsonWriterPipeline:
    """
    Writes items to JSON file incrementally as they're scraped.
    This ensures data is saved even if the scraper gets stuck or interrupted.
    """

    def open_spider(self, spider):
        """Called when spider opens - initialize the JSON file"""
        # Check format - only activate if JSON format is selected
        format_type = getattr(spider, 'output_format', 'json').lower()
        if format_type != 'json':
            self.file = None
            return  # Deactivate if not JSON format

        # Get output file from spider settings or use default
        output_file = getattr(spider, 'output_file', None)
        if not output_file:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"output_{timestamp}.json"

        # Force .json extension for consistency with other pipelines
        output_file = output_file.rsplit('.', 1)[0] + '.json'

        self.output_file = output_file
        self.file = open(self.output_file, 'w', encoding='utf-8')
        self.file.write('[\n')  # Start JSON array
        self.first_item = True
        self.item_count = 0

        spider.logger.info(f"📝 Incremental JSON writer initialized: {self.output_file}")

    def close_spider(self, spider):
        """Called when spider closes - finalize the JSON file"""
        if self.file is None:  # Pipeline was deactivated
            return

        self.file.write('\n]')  # Close JSON array
        self.file.close()
        spider.logger.info(f"✅ Saved {self.item_count} items to {self.output_file}")

    def process_item(self, item, spider):
        """Write each item to JSON immediately"""
        if self.file is None:  # Pipeline was deactivated
            return item

        # Add comma before item if not first
        if not self.first_item:
            self.file.write(',\n')
        else:
            self.first_item = False

        # Write item as JSON
        line = json.dumps(dict(item), ensure_ascii=False, indent=2)
        self.file.write('  ' + line.replace('\n', '\n  '))  # Indent for readability
        self.file.flush()  # Force write to disk immediately

        self.item_count += 1
        spider.logger.info(f"💾 Saved item #{self.item_count}: {item.get('name', 'Unknown')}")

        return item


class IncrementalCsvWriterPipeline:
    """
    Writes items to CSV file incrementally with UTF-8 BOM for Excel international character support.
    Only activates when spider.output_format == 'csv'.
    """

    def open_spider(self, spider):
        """Initialize CSV file with UTF-8 BOM and headers"""
        # Only activate if format is CSV
        format_type = getattr(spider, 'output_format', 'json').lower()
        if format_type != 'csv':
            self.file = None
            return

        # Get output file and change extension to .csv
        output_file = getattr(spider, 'output_file', None)
        if not output_file:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"output_{timestamp}.csv"

        output_file = output_file.rsplit('.', 1)[0] + '.csv'

        self.output_file = output_file
        # CRITICAL: utf-8-sig adds BOM for Excel international compatibility
        self.file = open(self.output_file, 'w', encoding='utf-8-sig', newline='')
        self.csv_writer = csv.DictWriter(
            self.file,
            fieldnames=['name', 'email', 'position', 'company_url', 'page_url'],
            extrasaction='ignore'
        )
        self.csv_writer.writeheader()
        self.file.flush()
        self.item_count = 0

        spider.logger.info(f"📝 Incremental CSV writer initialized: {self.output_file}")

    def close_spider(self, spider):
        """Close CSV file"""
        if self.file:
            self.file.close()
            spider.logger.info(f"✅ Saved {self.item_count} items to {self.output_file}")

    def process_item(self, item, spider):
        """Write each item to CSV immediately"""
        if not self.file:
            return item  # Not active for this format

        self.csv_writer.writerow(dict(item))
        self.file.flush()  # Immediate disk write

        self.item_count += 1
        spider.logger.info(f"💾 Saved CSV item #{self.item_count}: {item.get('name', 'Unknown')}")

        return item


class IncrementalExcelWriterPipeline:
    """
    Writes items to Excel (XLSX) file incrementally using load/append/save pattern.
    Only activates when spider.output_format == 'xlsx'.
    """

    def open_spider(self, spider):
        """Initialize Excel file with styled headers"""
        # Only activate if format is XLSX
        format_type = getattr(spider, 'output_format', 'json').lower()
        if format_type != 'xlsx':
            self.output_file = None
            return

        # Get output file and change extension to .xlsx
        output_file = getattr(spider, 'output_file', None)
        if not output_file:
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f"output_{timestamp}.xlsx"

        output_file = output_file.rsplit('.', 1)[0] + '.xlsx'

        self.output_file = output_file
        self.item_count = 0

        # Create workbook with styled headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scraped Data"

        # Write headers
        headers = ['Name', 'Email', 'Position', 'Company URL', 'Page URL']
        sheet.append(headers)

        # Style headers: bold + light blue background
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Set column widths
        sheet.column_dimensions['A'].width = 25  # Name
        sheet.column_dimensions['B'].width = 30  # Email
        sheet.column_dimensions['C'].width = 25  # Position
        sheet.column_dimensions['D'].width = 40  # Company URL
        sheet.column_dimensions['E'].width = 40  # Page URL

        workbook.save(self.output_file)

        spider.logger.info(f"📝 Incremental Excel writer initialized: {self.output_file}")

    def close_spider(self, spider):
        """Finalize Excel file"""
        if self.output_file:
            spider.logger.info(f"✅ Saved {self.item_count} items to {self.output_file}")

    def process_item(self, item, spider):
        """Write each item to Excel immediately"""
        if not self.output_file:
            return item  # Not active for this format

        try:
            # Load, append, save (incremental pattern)
            workbook = load_workbook(self.output_file)
            sheet = workbook.active

            row_data = [
                item.get('name', ''),
                item.get('email', ''),
                item.get('position', ''),
                item.get('company_url', ''),
                item.get('page_url', '')
            ]
            sheet.append(row_data)

            workbook.save(self.output_file)  # Immediate save

            self.item_count += 1
            spider.logger.info(f"💾 Saved Excel item #{self.item_count}: {item.get('name', 'Unknown')}")

        except PermissionError:
            spider.logger.error(f"❌ Cannot save Excel file - is it open in Excel? Close it and restart.")
            raise

        return item


class StateTrackerPipeline:
    """
    Tracks scraping progress in memory for logging purposes.
    Note: State tracking file output has been removed as it was unused.
    """

    def open_spider(self, spider):
        """Initialize state tracking (in-memory only)"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.state = {
            'start_time': timestamp,
            'start_url': getattr(spider, 'start_url', ''),
            'items_scraped': 0,
            'last_page': None,
            'status': 'running'
        }
        spider.logger.info(f"📊 State tracking initialized (in-memory)")

    def close_spider(self, spider):
        """Log final state"""
        from datetime import datetime
        self.state['end_time'] = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.state['status'] = 'completed'
        spider.logger.info(f"✅ Scraping completed: {self.state['items_scraped']} items scraped")

    def process_item(self, item, spider):
        """Update state with each item"""
        self.state['items_scraped'] += 1
        self.state['last_page'] = item.get('page_url', '')
        return item
